import requests
import pandas
import shutil
import os
import numpy as np
import sys
import arcpy


try:
    import openpyxl
except ImportError:
    arcpy.AddMessage("Error: Your Python installation is missing the openpyxl module. Please install the module before running this script.")
    sys.exit()

from openpyxl import load_workbook

#allows user to specify the subarea at runtime
subarea = str(arcpy.GetParameterAsText(0))

#copy the matching template into a unique file to serve
#as the matching sheet for the subarea
copy_name = r"C:\RECON\%s_Matching.xlsx" %(subarea)
if os.path.exists(r"C:\RECON\Matching_Template.xlsx"):
    shutil.copy(r"C:\RECON\Matching_Template.xlsx", copy_name)

#specifies the map document as the map in which the script is run, and
#the data frame as the first data frame listed
arcpy.env.workspace = r"C:\RECON"
mxd = arcpy.mapping.MapDocument("CURRENT")
arcpy.env.overwriteOutput = True
dataFrame = arcpy.mapping.ListDataFrames(mxd)[0]

#retrieves the pump station subarea, writes a definition query for that PSNUM,
#then sets the extent of the data frame to that queried layer
lyr = arcpy.mapping.Layer("Pump Station Subareas")
lyr.definitionQuery = "PSNUM = '%s'" %(subarea)
extent = lyr.getExtent(True)
dataFrame.extent = extent
arcpy.RefreshActiveView()

del lyr

#creates a field mappings container, then a list containing 6 field map objects
#for each of the fields required for the RECON process. the SHAPE.LEN field must
#be renamed because arcpy does not allow writing '.' in field names, which
#causes the writing of that field to fail
lyr = arcpy.mapping.Layer("SEWER.GravityMain")
fieldMappings = arcpy.FieldMappings()
mapsList = []
for n in range(0,6):
     mapsList.append(arcpy.FieldMap())
mapsList[0].addInputField(lyr,"MAX_NUM")
mapsList[1].addInputField(lyr,"LIFECYCLESTATUS")
mapsList[2].addInputField(lyr,"MATERIAL")
mapsList[3].addInputField(lyr,"LINING")
mapsList[4].addInputField(lyr,"DIAMETER")
mapsList[5].addInputField(lyr,"SHAPE.LEN")
fld_SHAPE_LEN = mapsList[5].outputField
fld_SHAPE_LEN.name = "LEN"
mapsList[5].outputField = fld_SHAPE_LEN

#each field map object is added to the field mappings container
for n in range(0,len(mapsList)):
    fieldMappings.addFieldMap(mapsList[n])

#the table to be copied is selected, queried, and converted to a .csv using the
#field mappings created above    
arcpy.management.SelectLayerByAttribute(lyr,"NEW_SELECTION", "MAX_NUM LIKE '" + subarea + "%'")
arcpy.conversion.TableToTable(lyr, "C:\RECON", subarea + "_GIS.csv", "MAX_NUM LIKE '" + subarea + "%'", fieldMappings)

#uses pandas library to sort the created .csv file ascendingly using MAX_NUM
df = pandas.read_csv("C:\\RECON\\" + subarea + "_GIS.csv")
sortedCSV = df.sort_values(by=["MAX_NUM"])

#loop through the sortedCSV pandas dataframe,
#pasting values into the respective cells in the
#new excel matching file. if any value in the dataframe
#is 'not a number' (NaN), the sheet is populated with a
#string '<Null>'.
book = load_workbook(copy_name)
sheet = book.get_sheet_by_name('Sheet1')
row = 2
column = 5
index = 0
series = 1
for rows in range(0, len(sortedCSV.index + 1)):
    for field in range(1,7):
        mycell = sheet.cell(row, column)
        if np.nan in [sortedCSV.iat[index,field]]:
            mycell.value = "<Null>"
        else:
            mycell.value = sortedCSV.iat[index,field]
        column += 1
    index += 1    
    column -= 6        
    row += 1

#save and close the book
book.save(copy_name)
book.close()

#create the url used to establish the server connection. api conventions are based on maximo JSON
#api for version 7.5.x
conn = "http://ocudrhpdmx01"
url = conn + "/maxrest/rest/mbo/locations"

#supply required headers
headers = {
    'Accept': "application/json",
    'Content-Type':"application/json"}

#supply search and sort parameters
params = {
    '_format':'json',
    'compact':1,
    'location':'~sw~%s' % subarea,
    'PLUSSFEATURECLASS':'GISSEWERGRAVITYMAIN',
    'orderby':'location',
    'status':'operating,planned'}

#create the variable for the get request
res = requests.get(url, headers = headers, params = params, auth=('136048','@Tahsil63'), verify = False)

#loop through the json returned in the response object 'res',
#pasting values into the respective cells in the
#new excel matching file.
book = load_workbook(copy_name)
sheet = book.get_sheet_by_name('Sheet1')
row = 2
column = 1
resList = res.json().get('LOCATIONSMboSet').get('LOCATIONS')
for i,k in enumerate(resList):
    if 'LOCATION' in resList[i]['Attributes'].keys():
        mycell = sheet.cell(row, column)
        mycell.value = (str(k['Attributes']['LOCATION']['content']))
        column += 1
    elif 'DESCRIPTION' in resList[i]['Attributes'].keys():
        mycell = sheet.cell(row, column)
        mycell.value =(str(k['Attributes']['DESCRIPTION']['content']))
        column -= 1
        row += 1
    else:
        continue

#save and close the book
book.save(copy_name)
book.close()

del url
del params
del res
del resList


url = conn + '/maxrest/rest/mbo/routes'

params = {
    '_format':'json',
    'compact':1,
    'description':"GRAVITY PM, SA%s" %subarea
    }

res = requests.get(url, headers = headers, params = params, auth=('136048','@Tahsil63'), verify = False)

resList = res.json().get('ROUTESMboSet').get('ROUTES')
routeNum = ''

#loops through the json returned in the new response object
#'res', finding the route number associated with the
#queried route
for i,k in enumerate(resList):
    if 'ROUTE' in resList[i]['Attributes'].keys():
        routeNum = k['Attributes']['ROUTE']['content']
    else:
        continue

del url
del params
del res
del resList

url = conn + '/maxrest/rest/mbo/route_stop'

params = {
    '_format':'json',
    'compact':1,
    '_orderby':'location',
    'route':routeNum,
    }

res = requests.get(url, headers = headers, params = params, auth=('136048','@Tahsil63'), verify = False)

resList = res.json().get('ROUTE_STOPMboSet').get('ROUTE_STOP')

book = load_workbook(copy_name)
sheet = book.get_sheet_by_name('Sheet1')
row = 2
column = 3

#loops through the json returned in the new response object
#'res', pasting the location number of each object in the route
#in the new excel matching file
for i,k in enumerate(resList):
    if 'LOCATION' in resList[i]['Attributes'].keys():
        mycell = sheet.cell(row, column)
        mycell.value = str(k['Attributes']['LOCATION']['content'])
        row += 1
    else:
        continue

book.save(copy_name)
book.close()

#print a message to the arcmap python console indicating
#completion.
arcpy.AddMessage("Done.")



        
                             



    




